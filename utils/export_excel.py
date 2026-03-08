"""
Excel (.xlsx) Dışa Aktarma Modülü
openpyxl kullanır — 3 sekme: Özet | GVK Tarifesi | Ücret Detay
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from utils.hesapla import HesaplaSonucu, format_tl


# ── Renk paleti ───────────────────────────────────────────────────────────
NAVY   = "1A2744"
GOLD   = "C49A2B"
WHITE  = "FFFFFF"
LGRAY  = "F5F1EB"
DGRAY  = "6B7280"
GREEN  = "27AE60"
RED    = "C0392B"
AMBER  = "FFF8E1"
LGREEN = "E8F5E9"

TL_FORMAT = '₺#,##0.00'

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=None, size=11) -> Font:
    return Font(bold=bold, color=color or "000000", size=size, name="Arial")

def _border(style="thin") -> Border:
    s = Side(border_style=style, color="E5DDC8")
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border() -> Border:
    s = Side(border_style="thin", color="E5DDC8")
    return Border(bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _sec_header(ws, row: int, text: str, max_col: int = 3):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _font(bold=True, color=WHITE, size=10)
    cell.fill = _fill(NAVY)
    cell.alignment = Alignment(vertical="center", indent=1)
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = _fill(NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    ws.row_dimensions[row].height = 18


def _data_row(ws, row: int, label: str, value, is_total=False, is_green=False,
              value_format=TL_FORMAT, note: str = ""):
    bg = LGRAY if row % 2 == 0 else WHITE
    if is_total:
        bg = NAVY

    label_cell = ws.cell(row=row, column=1, value=label)
    value_cell = ws.cell(row=row, column=2, value=value)
    note_cell  = ws.cell(row=row, column=3, value=note) if note else None

    for cell in [label_cell, value_cell]:
        cell.fill = _fill(bg)
        cell.border = _bottom_border()

    label_cell.font = _font(
        bold=is_total,
        color=WHITE if is_total else DGRAY,
        size=10
    )
    label_cell.alignment = Alignment(vertical="center", indent=1)

    value_cell.font = _font(
        bold=is_total,
        color=WHITE if is_total else (GREEN if is_green else NAVY),
        size=10
    )
    value_cell.alignment = _right()
    if isinstance(value, (int, float)):
        value_cell.number_format = value_format

    if note_cell:
        note_cell.fill = _fill(bg)
        note_cell.font = _font(size=9, color=DGRAY)
        note_cell.alignment = Alignment(vertical="center", indent=1)

    ws.row_dimensions[row].height = 16


# ── Sayfa 1: Özet ─────────────────────────────────────────────────────────

def _build_ozet(wb: Workbook, sonuc: HesaplaSonucu, year: int,
                gider_yontemi: str, params: dict):
    ws = wb.active
    ws.title = "Özet"
    ws.sheet_view.showGridLines = False

    # Başlık bandı
    for r in range(1, 5):
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = _fill(NAVY)
    ws.merge_cells("A1:C3")
    title = ws["A1"]
    title.value = f"GMSİ VERGİ HESAPLAMA ÖZETİ — {year}"
    title.font = _font(bold=True, color=WHITE, size=14)
    title.alignment = _center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A4:C4")
    sub = ws["A4"]
    sub.value = (
        f"Hesaplama Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  "
        f"GİB Parametreleri: gib.gov.tr  |  Bilgilendirme amaçlıdır"
    )
    sub.font = _font(color="A0AEC0", size=9)
    sub.alignment = _center()
    ws.row_dimensions[4].height = 14

    r = 6

    # GMSİ Gelirleri
    _sec_header(ws, r, "GMSİ GELİRLERİ"); r += 1
    _data_row(ws, r, "Konut Brüt Kira Geliri",       sonuc.konut_brut);     r += 1
    _data_row(ws, r, "İşyeri Kira Geliri (Stopajlı Brüt)", sonuc.isyeri_s_brut); r += 1
    _data_row(ws, r, "İşyeri Kira Geliri (Stopajsız)",sonuc.isyeri_n_brut); r += 1

    # İstisna
    _sec_header(ws, r, "İSTİSNA"); r += 1
    _data_row(ws, r, f"Mesken İstisnası ({year})",
              -sonuc.istisna if sonuc.istisna else 0,
              is_green=sonuc.istisna > 0,
              note="GVK Md.21" if sonuc.istisna > 0 else "Uygulanmadı"); r += 1
    _data_row(ws, r, "Beyana Tabi GMSİ Toplamı",     sonuc.toplam_gmsi_beyan, is_total=True); r += 1

    # Gider
    gider_lbl = (
        f"Gider — Götürü (%{int(params['goturu_oran']*100)})"
        if gider_yontemi == "goturu"
        else f"Gider — Gerçek Gider{sonuc.iktisap_acik}"
    )
    _sec_header(ws, r, "GİDER"); r += 1
    _data_row(ws, r, gider_lbl, -sonuc.gmsi_gider, is_green=True); r += 1
    _data_row(ws, r, "GMSİ Vergi Matrahı",            sonuc.gmsi_matrah, is_total=True); r += 1

    # Ücret
    _sec_header(ws, r, "ÜCRET GELİRİ"); r += 1
    _data_row(ws, r, "Toplam Brüt Ücret (12 ay)",     sonuc.toplam_brut_ucret); r += 1
    _data_row(ws, r, "Beyana Dahil Ücret Matrahı",    sonuc.ucret_matrah,
              note="Beyan zorunlu" if sonuc.ucret_beyan_zorunlu else "Beyan gerekmez"); r += 1

    # Vergi
    _sec_header(ws, r, "VERGİ HESABI"); r += 1
    _data_row(ws, r, "Toplam Vergi Matrahı",           sonuc.toplam_matrah); r += 1
    _data_row(ws, r, "Hesaplanan Gelir Vergisi",       sonuc.hes_ver); r += 1
    _data_row(ws, r, "Mahsup Edilen Stopaj",          -sonuc.mahsup, is_green=True); r += 1
    _data_row(ws, r, "ÖDENECEK GELİR VERGİSİ",        sonuc.odeme, is_total=True); r += 1
    if sonuc.iade > 0:
        _data_row(ws, r, "İADE EDİLECEK VERGİ",       sonuc.iade, is_green=True); r += 1
    _data_row(ws, r, "Damga Vergisi (yaklaşık)",       sonuc.damga); r += 1

    # Taksit
    _sec_header(ws, r, "TAKSİT PLANI"); r += 1
    _data_row(ws, r, f"1. Taksit + Damga — 31 Mart {year+1}",
              sonuc.taksit1 + sonuc.damga,
              note="Banka / PTT / Dijital Vergi Dairesi"); r += 1
    _data_row(ws, r, f"2. Taksit — 31 Temmuz {year+1}",
              sonuc.taksit2); r += 1

    # Uyarılar
    r += 1
    _sec_header(ws, r, "UYARILAR"); r += 1
    for w in sonuc.warns:
        cell = ws.cell(row=r, column=1, value=w)
        cell.font = _font(size=9, color=DGRAY)
        cell.alignment = Alignment(wrap_text=True, indent=1)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 24
        r += 1

    # Yasal uyarı
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cell = ws.cell(row=r, column=1,
                   value="⚠️ Bu hesaplama bilgilendirme amaçlıdır. "
                         "Kesin vergi yükümlülüğünüz için YMM/SMMM'ye danışınız.")
    cell.font = _font(size=9, color=RED)
    cell.fill = _fill("FFF8E1")
    cell.alignment = Alignment(wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 20

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 36


# ── Sayfa 2: GVK Tarifesi ─────────────────────────────────────────────────

def _build_tarife(wb: Workbook, sonuc: HesaplaSonucu, year: int, params: dict):
    ws = wb.create_sheet("GVK Tarifesi")
    ws.sheet_view.showGridLines = False

    headers = ["Alt Sınır (TL)", "Üst Sınır (TL)", "Oran (%)", "Taban Vergi (TL)"]
    col_ws  = [22, 22, 14, 22]

    # Başlık
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"GELİR VERGİSİ TARİFESİ — ÜCRET DIŞI — {year} YILI"
    t.font = _font(bold=True, color=WHITE, size=12)
    t.fill = _fill(NAVY)
    t.alignment = _center()
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = "Kaynak: GİB — gib.gov.tr | GVK Md. 103"
    s.font = _font(size=9, color=DGRAY)
    s.alignment = _center()
    ws.row_dimensions[2].height = 14

    # Sütun başlıkları
    for ci, (h, w) in enumerate(zip(headers, col_ws), 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = _font(bold=True, color=WHITE, size=10)
        c.fill = _fill(NAVY)
        c.alignment = _center()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 18

    # Tarife satırları
    for ri, dilim in enumerate(params["tarife_ucret_disi"], 5):
        bg = LGRAY if ri % 2 == 0 else WHITE
        values = [
            dilim["alt_sinir"],
            dilim["ust_sinir"] if dilim["ust_sinir"] else "Sınırsız",
            dilim["oran"] * 100,
            dilim["taban"],
        ]
        fmts = [TL_FORMAT, TL_FORMAT, "0.0", TL_FORMAT]
        for ci, (val, fmt) in enumerate(zip(values, fmts), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _font(size=10)
            c.fill = _fill(bg)
            c.alignment = _right() if ci > 1 else Alignment(indent=1, vertical="center")
            c.border = _bottom_border()
            if isinstance(val, (int, float)):
                c.number_format = fmt
        ws.row_dimensions[ri].height = 16

    # Özet
    r = len(params["tarife_ucret_disi"]) + 7
    _sec_header(ws, r, "HESAPLAMA ÖZETI", max_col=4); r += 1
    for lbl, val in [
        ("Toplam Vergi Matrahı",   sonuc.toplam_matrah),
        ("Hesaplanan Gelir Vergisi", sonuc.hes_ver),
        ("Efektif Vergi Oranı",
         sonuc.hes_ver / sonuc.toplam_matrah if sonuc.toplam_matrah > 0 else 0),
    ]:
        c1 = ws.cell(row=r, column=1, value=lbl)
        c2 = ws.cell(row=r, column=2, value=val)
        for c in [c1, c2, ws.cell(row=r, column=3), ws.cell(row=r, column=4)]:
            c.fill = _fill(LGRAY)
            c.border = _bottom_border()
        c1.font = _font(size=10)
        c1.alignment = Alignment(indent=1, vertical="center")
        c2.font = _font(bold=True, size=10, color=NAVY)
        c2.alignment = _right()
        c2.number_format = "0.00%" if lbl == "Efektif Vergi Oranı" else TL_FORMAT
        ws.row_dimensions[r].height = 16
        r += 1


# ── Sayfa 3: Ücret Detay ──────────────────────────────────────────────────

def _build_ucret(wb: Workbook, sonuc: HesaplaSonucu, params: dict,
                 isverenler: list):
    ws = wb.create_sheet("Ücret Detay")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "ÜCRET GELİRİ DETAYI"
    t.font = _font(bold=True, color=WHITE, size=12)
    t.fill = _fill(NAVY)
    t.alignment = _center()
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:E2")
    info = ws["A2"]
    info.value = (
        f"Beyan Eşiği (2.+ İşveren): {format_tl(params['ucret_beyan_esigi'])}  |  "
        f"Durum: {'⚠️ BEYAN ZORUNLU' if sonuc.ucret_beyan_zorunlu else '✓ Beyan Gerekmiyor'}"
    )
    info.font = _font(
        size=10,
        color=RED if sonuc.ucret_beyan_zorunlu else GREEN
    )
    info.alignment = _center()
    ws.row_dimensions[2].height = 18

    headers = ["İşveren Adı", "Sıra", "12 Ay Brüt Ücret (TL)", "Kesilen Stopaj (TL)", "Not"]
    col_ws  = [30, 18, 26, 26, 20]
    for ci, (h, w) in enumerate(zip(headers, col_ws), 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = _font(bold=True, color=WHITE, size=10)
        c.fill = _fill(NAVY)
        c.alignment = _center()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 18

    for ri, iv in enumerate(isverenler, 5):
        bg = LGRAY if ri % 2 == 0 else WHITE
        vals = [
            iv.ad or "(İsimsiz)",
            "1. İşveren" if iv.birinci else "2.+ İşveren",
            iv.brut_yillik,
            iv.stopaj,
            "En yüksek ücret" if iv.birinci else "Eşik kontrolüne tabi",
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _font(size=10)
            c.fill = _fill(bg)
            c.border = _bottom_border()
            c.alignment = _right() if ci in (3, 4) else Alignment(indent=1, vertical="center")
            if ci in (3, 4):
                c.number_format = TL_FORMAT
        ws.row_dimensions[ri].height = 16

    # Toplam
    tr = len(isverenler) + 6
    ws.cell(row=tr, column=1, value="TOPLAM").font = _font(bold=True, color=WHITE, size=10)
    ws.cell(row=tr, column=3, value=sonuc.toplam_brut_ucret).font = _font(bold=True, color=WHITE, size=10)
    ws.cell(row=tr, column=4, value=sonuc.toplam_ucret_stopaj).font = _font(bold=True, color=WHITE, size=10)
    for ci in range(1, 6):
        c = ws.cell(row=tr, column=ci)
        c.fill = _fill(NAVY)
        c.border = _bottom_border()
        if ci in (3, 4):
            c.alignment = _right()
            c.number_format = TL_FORMAT


# ── Ana export fonksiyonu ─────────────────────────────────────────────────

def export_excel(
    sonuc: HesaplaSonucu,
    year: int,
    gider_yontemi: str,
    params: dict,
    isverenler: list,
    ucret_var: bool,
) -> bytes:
    """
    Hesaplama sonuçlarını Excel formatına dönüştürür.
    Returns: bytes — doğrudan HTTP response olarak gönderilebilir.
    """
    wb = Workbook()

    _build_ozet(wb, sonuc, year, gider_yontemi, params)
    _build_tarife(wb, sonuc, year, params)
    if ucret_var and isverenler:
        _build_ucret(wb, sonuc, params, isverenler)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
